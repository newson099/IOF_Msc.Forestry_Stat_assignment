"""
Q14. Contributing factors for participation in BZ conservation activities
     -- specifically (a) User Group Mobilization and (b) Anti-poaching
     Operation -- with respect to socio-economic variables, using
     Ordered Probit Regression (appropriate because both dependent
     variables are ordinal: 1 = Low, 2 = Average, 3 = High).

"""

import os
import numpy as np
import pandas as pd
from statsmodels.miscmodels.ordinal_model import OrderedModel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_PATH = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\spss.csv"
OUTPUT_DIR = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\14"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "q14_results.xlsx")

ALPHA = 0.05

df = pd.read_csv(DATA_PATH, encoding="utf-8-sig")
df["District"] = df["District"].replace("2+C182:BC182", "2").astype(int)

df["District_lbl"] = df["District"].map({1: "Bara", 2: "Parsa"})
df["Sex_lbl"] = df["HH_Head_sex"].map({1: "Female", 2: "Male"})
df["Ethnic_lbl"] = df["Ehnicity"].map({1: "Dalit", 2: "Indigenous", 3: "Madhesi", 4: "Brahmin_Chhetri"})
df["Eco_class_lbl"] = df["Eco_class"].map({1: "Poor", 2: "Middle", 3: "Rich"})
df["Occupation_lbl"] = df["Main_Occupation"].map({1: "Agriculture", 2: "Service", 3: "Business"})

# Dummy-code categorical predictors (drop_first gives a clean reference category)
X = pd.get_dummies(
    df[["Age", "Edu.Hh", "Family_size", "LSU", "Land_holding_ropani", "Dis_from",
        "District_lbl", "Sex_lbl", "Ethnic_lbl", "Eco_class_lbl", "Occupation_lbl"]],
    columns=["District_lbl", "Sex_lbl", "Ethnic_lbl", "Eco_class_lbl", "Occupation_lbl"],
    drop_first=True,
).astype(float)

# Readable predictor names for the output tables
readable_names = {
    "Age": "Age (years)",
    "Edu.Hh": "Education (years)",
    "Family_size": "Family Size",
    "LSU": "Livestock Standard Unit",
    "Land_holding_ropani": "Land Holding (ropani)",
    "Dis_from": "Distance from BZ (min)",
    "District_lbl_Parsa": "District: Parsa (vs Bara)",
    "Sex_lbl_Male": "Sex: Male (vs Female)",
    "Ethnic_lbl_Dalit": "Ethnicity: Dalit (vs Brahmin/Chhetri)",
    "Ethnic_lbl_Indigenous": "Ethnicity: Indigenous (vs Brahmin/Chhetri)",
    "Ethnic_lbl_Madhesi": "Ethnicity: Madhesi (vs Brahmin/Chhetri)",
    "Eco_class_lbl_Poor": "Economic Class: Poor (vs Middle)",
    "Eco_class_lbl_Rich": "Economic Class: Rich (vs Middle)",
    "Occupation_lbl_Service": "Occupation: Service (vs Agriculture)",
    "Occupation_lbl_Business": "Occupation: Business (vs Agriculture)",
}


def run_ordered_probit(y, activity_name):
    """Fit an ordered probit model and return a clean coefficient table
    plus a model-fit summary table."""
    model = OrderedModel(y, X, distr="probit")
    result = model.fit(method="bfgs", disp=False)
    print(result.summary())

    n_predictors = X.shape[1]
    coef_names = list(X.columns) + list(result.params.index[n_predictors:])  # predictors + thresholds

    coef_table = pd.DataFrame({
        "Predictor": [readable_names.get(n, n) for n in X.columns] + list(result.params.index[n_predictors:]),
        "Coefficient": result.params.values.round(4),
        "Std_Error": result.bse.values.round(4),
        "z_stat": result.tvalues.round(3),
        "p_value": result.pvalues.round(4),
        "Significant_5pct": result.pvalues < ALPHA,
    })

    category_counts = y.value_counts()
    category_props = category_counts / len(y)
    null_llf = float((category_counts * np.log(category_props)).sum())
    pseudo_r2 = 1 - (result.llf / null_llf)

    model_summary = pd.DataFrame({
        "Statistic": ["N", "Log-Likelihood", "LL-Null", "Pseudo R-squared (McFadden)",
                      "LR chi-square", "LR p-value"],
        "Value": [int(result.nobs), round(result.llf, 3), round(null_llf, 3),
                  round(pseudo_r2, 4), round(2 * (result.llf - null_llf), 3),
                  "see coefficients for individual significance"],
    })

    print(f"\n{activity_name}: Pseudo R2 = {pseudo_r2:.4f}\n")
    return coef_table, model_summary


print("=" * 70)
print("(a) ORDERED PROBIT -- Participation: User Group Mobilization")
print("=" * 70)
ug_coef, ug_summary = run_ordered_probit(df["Part_UG_mobilization"], "User Group Mobilization")

print("=" * 70)
print("(b) ORDERED PROBIT -- Participation: Anti-poaching Operation")
print("=" * 70)
poach_coef, poach_summary = run_ordered_probit(df["Part_anti_poaching"], "Anti-poaching Operation")


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SIG_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4


def write_table(ws, df_table, start_row=1, title=None, highlight_sig_col=None, first_col_left=False):
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 2

    header_row = row
    for c, col_name in enumerate(df_table.columns, start=1):
        ws.cell(row=header_row, column=c, value=col_name)
    style_header_row(ws, header_row, len(df_table.columns))
    row += 1

    for _, data_row in df_table.iterrows():
        for c, col_name in enumerate(df_table.columns, start=1):
            cell = ws.cell(row=row, column=c, value=data_row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = LEFT if (first_col_left and c == 1) else CENTER
            if highlight_sig_col is not None and col_name == highlight_sig_col and bool(data_row[col_name]):
                for cc in range(1, len(df_table.columns) + 1):
                    ws.cell(row=row, column=cc).fill = SIG_FILL
        row += 1
    return row + 2


with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="UG_Mobilization")
    pd.DataFrame().to_excel(writer, sheet_name="Anti_Poaching")
    wb = writer.book

    ws1 = wb["UG_Mobilization"]
    next_row = write_table(
        ws1, ug_coef, start_row=1,
        title="Ordered Probit: Participation in User Group Mobilization (green = significant at 5%)",
        highlight_sig_col="Significant_5pct", first_col_left=True,
    )
    write_table(ws1, ug_summary, start_row=next_row, title="Model Fit Summary", first_col_left=True)
    autofit_columns(ws1)
    ws1.freeze_panes = "A2"

    ws2 = wb["Anti_Poaching"]
    next_row = write_table(
        ws2, poach_coef, start_row=1,
        title="Ordered Probit: Participation in Anti-poaching Operation (green = significant at 5%)",
        highlight_sig_col="Significant_5pct", first_col_left=True,
    )
    write_table(ws2, poach_summary, start_row=next_row, title="Model Fit Summary", first_col_left=True)
    autofit_columns(ws2)
    ws2.freeze_panes = "A2"

print(f"\nAll Q14 results saved (formatted) to: {OUTPUT_FILE}")
