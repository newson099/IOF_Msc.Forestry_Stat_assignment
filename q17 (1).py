"""
Q17. Influencing factors for the loss of number of domestic animals
     (count data: cows + buffaloes + goats lost per household), using
     Poisson Regression -- appropriate for count outcomes (non-negative
     integers), unlike OLS which assumes a continuous, unbounded response.

     Dependent variable: Total_animals_lost = Cow_no + Buf.no + Gt.no
     Independent variables: same socio-economic set used in Q11/Q13-Q16

"""

import os
import numpy as np
import pandas as pd
import statsmodels.api as sm
from statsmodels.formula.api import glm
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_PATH = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\spss.csv"
OUTPUT_DIR = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\17"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "q17_results.xlsx")

ALPHA = 0.05

df = pd.read_csv(DATA_PATH, encoding="utf-8-sig")
df["District"] = df["District"].replace("2+C182:BC182", "2").astype(int)

df["Total_animals_lost"] = df["Cow_no"] + df["Buf.no"] + df["Gt.no"]

print("Total animals lost -- descriptive check (mean vs variance):")
print(f"  Mean     = {df['Total_animals_lost'].mean():.3f}")
print(f"  Variance = {df['Total_animals_lost'].var():.3f}")
print("  (Poisson assumes Mean = Variance; a ratio noticeably > 1 suggests mild overdispersion,")
print("   in which case a Negative Binomial model would be a useful robustness check.)\n")

print("Distribution of animals lost:")
print(df["Total_animals_lost"].value_counts().sort_index(), "\n")

df["District_lbl"] = df["District"].map({1: "Bara", 2: "Parsa"})
df["Sex_lbl"] = df["HH_Head_sex"].map({1: "Female", 2: "Male"})
df["Ethnic_lbl"] = df["Ehnicity"].map({1: "Dalit", 2: "Indigenous", 3: "Madhesi", 4: "Brahmin_Chhetri"})
df["Eco_class_lbl"] = df["Eco_class"].map({1: "Poor", 2: "Middle", 3: "Rich"})
df["Occupation_lbl"] = df["Main_Occupation"].map({1: "Agriculture", 2: "Service", 3: "Business"})

# Fit the Poisson regression (GLM with Poisson family, log link)
formula = (
    "Total_animals_lost ~ Age + Q('Edu.Hh') + Family_size + LSU + Land_holding_ropani + Dis_from "
    "+ C(District_lbl) + C(Sex_lbl) + C(Ethnic_lbl) + C(Eco_class_lbl) + C(Occupation_lbl)"
)
model = glm(formula, data=df, family=sm.families.Poisson()).fit()
print(model.summary())

readable_names = {
    "Intercept": "Intercept",
    "C(District_lbl)[T.Parsa]": "District: Parsa (vs Bara)",
    "C(Sex_lbl)[T.Male]": "Sex: Male (vs Female)",
    "C(Ethnic_lbl)[T.Dalit]": "Ethnicity: Dalit (vs Brahmin/Chhetri)",
    "C(Ethnic_lbl)[T.Indigenous]": "Ethnicity: Indigenous (vs Brahmin/Chhetri)",
    "C(Ethnic_lbl)[T.Madhesi]": "Ethnicity: Madhesi (vs Brahmin/Chhetri)",
    "C(Eco_class_lbl)[T.Poor]": "Economic Class: Poor (vs Middle)",
    "C(Eco_class_lbl)[T.Rich]": "Economic Class: Rich (vs Middle)",
    "C(Occupation_lbl)[T.Business]": "Occupation: Business (vs Agriculture)",
    "C(Occupation_lbl)[T.Service]": "Occupation: Service (vs Agriculture)",
    "Age": "Age (years)",
    "Q('Edu.Hh')": "Education (years)",
    "Family_size": "Family Size",
    "LSU": "Livestock Standard Unit",
    "Land_holding_ropani": "Land Holding (ropani)",
    "Dis_from": "Distance from BZ (min)",
}

coef_table = pd.DataFrame({
    "Predictor": [readable_names.get(t, t) for t in model.params.index],
    "Coefficient": model.params.values.round(4),
    "Std_Error": model.bse.values.round(4),
    "z_stat": model.tvalues.round(3),
    "p_value": model.pvalues.round(4),
    "IRR (exp(coef))": np.exp(model.params.values).round(3),
    "IRR_CI_lower": np.exp(model.conf_int()[0].values).round(3),
    "IRR_CI_upper": np.exp(model.conf_int()[1].values).round(3),
    "Significant_5pct": model.pvalues.values < ALPHA,
})
print("\n", coef_table.to_string(index=False))

dispersion = model.pearson_chi2 / model.df_resid
print(f"\nDispersion statistic (Pearson chi2 / df) = {dispersion:.3f}")
if dispersion > 1.5:
    print("-> Notable overdispersion detected; a Negative Binomial model is recommended")
    print("   as a robustness check, since Poisson SEs may be understated.\n")
else:
    print("-> Dispersion is close to 1; the Poisson assumption (Mean = Variance) is reasonably met.\n")

model_summary = pd.DataFrame({
    "Statistic": ["N", "Log-Likelihood", "Deviance", "Pearson chi2", "Dispersion (Pearson chi2/df)",
                  "AIC", "BIC"],
    "Value": [int(model.nobs), round(model.llf, 3), round(model.deviance, 3),
              round(model.pearson_chi2, 3), round(dispersion, 3),
              round(model.aic, 3), round(model.bic, 3)],
})
print("Model Summary:")
print(model_summary.to_string(index=False))


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SIG_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4


def write_table(ws, df_table, start_row=1, title=None, highlight_sig_col=None, first_col_left=False):
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 2

    header_row = row
    for c, col_name in enumerate(df_table.columns, start=1):
        ws.cell(row=header_row, column=c, value=col_name)
    style_header_row(ws, header_row, len(df_table.columns))
    row += 1

    for _, data_row in df_table.iterrows():
        for c, col_name in enumerate(df_table.columns, start=1):
            cell = ws.cell(row=row, column=c, value=data_row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = LEFT if (first_col_left and c == 1) else CENTER
            if highlight_sig_col is not None and col_name == highlight_sig_col and bool(data_row[col_name]):
                for cc in range(1, len(df_table.columns) + 1):
                    ws.cell(row=row, column=cc).fill = SIG_FILL
        row += 1
    return row + 2


with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Q17_Poisson_Regression")
    wb = writer.book
    ws = wb["Q17_Poisson_Regression"]

    next_row = write_table(
        ws, coef_table, start_row=1,
        title="Poisson Regression: Determinants of Domestic Animals Lost (green = significant at 5%)",
        highlight_sig_col="Significant_5pct", first_col_left=True,
    )
    write_table(ws, model_summary, start_row=next_row, title="Model Fit Summary", first_col_left=True)

    autofit_columns(ws)
    ws.freeze_panes = "A2"

print(f"\nAll Q17 results saved (formatted) to: {OUTPUT_FILE}")
