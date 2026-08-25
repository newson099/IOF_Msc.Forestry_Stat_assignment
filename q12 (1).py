"""
Q12. Regression of Total Forest Product Income (dependent) on Economic
     Class (independent), using two dummy variables with a suitable
     reference category.

     Reference category chosen = "Poor" (the lowest economic class), so
     the two dummies represent Middle-vs-Poor and Rich-vs-Poor -- this is
     the most natural baseline for interpreting how income changes as
     households move up the economic ladder.
"""

import os
import pandas as pd
import statsmodels.api as sm
from statsmodels.formula.api import ols
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_PATH = r"c:\Users\anews\OneDrive\Desktop\Stat_Newson_assign\Py_CodeForStat_assignment\Output\Data.csv"
OUTPUT_DIR = r"c:\Users\anews\OneDrive\Desktop\Stat_Newson_assign\Py_CodeForStat_assignment\Output"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "q12_results.xlsx")

df = pd.read_csv(DATA_PATH, encoding="utf-8-sig")
df["District"] = df["District"].replace("2+C182:BC182", "2").astype(int)

df["Forest_income"] = (
    df["Fuel_qty"] * df["Fuel_price"]
    + df["Grass_qty"] * df["Grass_price"]
    + df["Leaf_qty"] * df["Leaf_pric"]
)

df["Eco_class_lbl"] = df["Eco_class"].map({1: "Poor", 2: "Middle", 3: "Rich"})

# dummy variables manually (reference = Poor) ----
df["D_Middle"] = (df["Eco_class_lbl"] == "Middle").astype(int)  # 1 if Middle, else 0
df["D_Rich"] = (df["Eco_class_lbl"] == "Rich").astype(int)      # 1 if Rich, else 0

# fitting regression
model = ols("Forest_income ~ D_Middle + D_Rich", data=df).fit()
print(model.summary())

group_means = df.groupby("Eco_class_lbl")["Forest_income"].agg(["count", "mean", "std"]).round(2)
group_means.columns = ["N", "Mean", "Std.Dev"]
group_means = group_means.reindex(["Poor", "Middle", "Rich"])
print("\nGroup means (Forest Income by Economic Class):")
print(group_means, "\n")

coef_table = pd.DataFrame({
    "Term": ["Intercept (Poor - reference)", "D_Middle (Middle vs Poor)", "D_Rich (Rich vs Poor)"],
    "Coefficient": model.params.values.round(3),
    "Std_Error": model.bse.values.round(3),
    "t_stat": model.tvalues.values.round(3),
    "p_value": model.pvalues.values.round(4),
    "Significant_5pct": (model.pvalues.values < 0.05),
})

model_summary = pd.DataFrame({
    "Statistic": ["N", "R-squared", "Adj. R-squared", "F-statistic", "F p-value"],
    "Value": [int(model.nobs), round(model.rsquared, 4), round(model.rsquared_adj, 4),
              round(model.fvalue, 3), round(model.f_pvalue, 4)],
})

print(coef_table.to_string(index=False))
print(f"\nRegression equation: Forest_income = {model.params.iloc[0]:.2f} "
      f"+ ({model.params.iloc[1]:.2f}) x D_Middle + ({model.params.iloc[2]:.2f}) x D_Rich")


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SIG_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)
CENTER = Alignment(horizontal="center", vertical="center")


def style_header_row(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4


def write_table(ws, df_table, start_row=1, title=None, highlight_sig_col=None):
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 2

    header_row = row
    for c, col_name in enumerate(df_table.columns, start=1):
        ws.cell(row=header_row, column=c, value=col_name)
    style_header_row(ws, header_row, len(df_table.columns))
    row += 1

    for _, data_row in df_table.iterrows():
        for c, col_name in enumerate(df_table.columns, start=1):
            cell = ws.cell(row=row, column=c, value=data_row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if highlight_sig_col is not None and col_name == highlight_sig_col and bool(data_row[col_name]):
                for cc in range(1, len(df_table.columns) + 1):
                    ws.cell(row=row, column=cc).fill = SIG_FILL
        row += 1
    return row + 2  # next free row, with a blank line gap


with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Q12_Regression_Result")
    wb = writer.book
    ws = wb["Q12_Regression_Result"]

    next_row = write_table(
        ws, group_means.reset_index().rename(columns={"index": "Economic Class"}),
        start_row=1, title="Group Means: Forest Income by Economic Class"
    )
    next_row = write_table(
        ws, coef_table, start_row=next_row,
        title="Regression Coefficients (Reference category = Poor)",
        highlight_sig_col="Significant_5pct",
    )
    write_table(
        ws, model_summary, start_row=next_row, title="Model Summary"
    )

    autofit_columns(ws)
    ws.freeze_panes = "A2"

print(f"\nAll Q12 results saved (formatted) to: {OUTPUT_FILE}")
