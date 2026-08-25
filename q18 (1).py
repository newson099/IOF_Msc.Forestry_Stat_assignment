"""
Q18. Principal Component Analysis (PCA): reduce the six continuous
     household-level characteristics into a smaller number of components,
     and identify the key underlying dimensions these represent (which,
     as predictors used throughout this assessment, capture the aspects
     of a household's socio-economic profile most relevant to explaining
     its loss/vulnerability outcomes).

     Variables: Age, Education, Family Size, LSU, Land Holding, Distance
                from BZ -- the same six continuous socio-economic
                variables used as predictors in Q11, Q13-Q17.

"""

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from factor_analyzer.factor_analyzer import calculate_kmo, calculate_bartlett_sphericity
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_PATH = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\spss.csv"
OUTPUT_DIR = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\18"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "q18_results.xlsx")
SCREE_PLOT = os.path.join(OUTPUT_DIR, "q18_scree_plot.png")

df = pd.read_csv(DATA_PATH, encoding="utf-8-sig")
df["District"] = df["District"].replace("2+C182:BC182", "2").astype(int)

variables = {
    "Age": "Age (years)",
    "Edu.Hh": "Education (years)",
    "Family_size": "Family Size",
    "LSU": "Livestock Standard Unit",
    "Land_holding_ropani": "Land Holding (ropani)",
    "Dis_from": "Distance from BZ (min)",
}
X_raw = df[list(variables.keys())].rename(columns=variables)

# Check suitability of the data for PCA ----
kmo_all, kmo_model = calculate_kmo(X_raw)
chi_square_value, p_value = calculate_bartlett_sphericity(X_raw)

print("PCA Suitability Checks")
print(f"  Kaiser-Meyer-Olkin (KMO) measure of sampling adequacy = {kmo_model:.3f}")
print("    (KMO > 0.6 is generally considered acceptable for PCA)")
print(f"  Bartlett's Test of Sphericity: chi-square = {chi_square_value:.3f}, p-value = {p_value:.4f}")
print("    (H0: variables are uncorrelated -- a significant result supports using PCA)\n")

# standardize the variables (mean=0, sd=1) before PCA 
# Standardization is essential here because the 6 variables are on very
# different scales (years, count, ropani, minutes) without it, Land
# Holding and Distance would dominate the components purely due to scale.
scaler = StandardScaler()
X_std = scaler.fit_transform(X_raw)

# Fit PCA (keep all components initially to inspect eigenvalues) 
pca = PCA()
pca.fit(X_std)

eigenvalues = pca.explained_variance_
explained_var_pct = pca.explained_variance_ratio_ * 100
cumulative_var_pct = np.cumsum(explained_var_pct)

eigen_table = pd.DataFrame({
    "Component": [f"PC{i+1}" for i in range(len(eigenvalues))],
    "Eigenvalue": eigenvalues.round(3),
    "% of Variance": explained_var_pct.round(2),
    "Cumulative %": cumulative_var_pct.round(2),
})
print("Eigenvalues and Variance Explained:")
print(eigen_table.to_string(index=False))

# Decide number of components to retain (Kaiser criterion: eigenvalue > 1) 
n_components = int((eigenvalues > 1).sum())
print(f"\nKaiser criterion (eigenvalue > 1) suggests retaining {n_components} component(s).\n")

# Component loadings (correlation between each variable and each retained PC)
loadings = pca.components_[:n_components].T * np.sqrt(eigenvalues[:n_components])
loadings_table = pd.DataFrame(
    loadings,
    index=X_raw.columns,
    columns=[f"PC{i+1}" for i in range(n_components)],
).round(3)
print("Component Loadings (retained components):")
print(loadings_table.to_string())

# Scree plot
plt.figure(figsize=(7, 5))
plt.plot(range(1, len(eigenvalues) + 1), eigenvalues, marker="o", color="steelblue", linewidth=2)
plt.axhline(y=1, color="red", linestyle="--", label="Eigenvalue = 1 (Kaiser cutoff)")
plt.xlabel("Component Number")
plt.ylabel("Eigenvalue")
plt.title("Scree Plot: PCA of Six Household Characteristics")
plt.xticks(range(1, len(eigenvalues) + 1))
plt.legend()
plt.tight_layout()
plt.savefig(SCREE_PLOT, dpi=150)
plt.close()
print(f"\nScree plot saved to: {SCREE_PLOT}")

suitability_table = pd.DataFrame([{
    "KMO_Overall": round(kmo_model, 3),
    "Bartlett_chi_square": round(chi_square_value, 3),
    "Bartlett_p_value": round(p_value, 4),
    "N_components_retained (eigenvalue>1)": n_components,
    "Cumulative_Variance_Explained_%": round(cumulative_var_pct[n_components - 1], 2),
}])

#output
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4


def write_table(ws, df_table, start_row=1, title=None, first_col_left=False, index_as_col=False):
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 2

    table = df_table.reset_index() if index_as_col else df_table
    header_row = row
    for c, col_name in enumerate(table.columns, start=1):
        ws.cell(row=header_row, column=c, value=str(col_name))
    style_header_row(ws, header_row, len(table.columns))
    row += 1

    for _, data_row in table.iterrows():
        for c, col_name in enumerate(table.columns, start=1):
            cell = ws.cell(row=row, column=c, value=data_row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = LEFT if (first_col_left and c == 1) else CENTER
        row += 1
    return row + 2


with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Q18_PCA")
    wb = writer.book
    ws = wb["Q18_PCA"]

    next_row = write_table(ws, suitability_table, start_row=1, title="PCA Suitability Checks (KMO & Bartlett's Test)")
    next_row = write_table(ws, eigen_table, start_row=next_row, title="Eigenvalues and Variance Explained")
    write_table(
        ws, loadings_table, start_row=next_row,
        title="Component Loadings (retained components only)",
        first_col_left=True, index_as_col=True,
    )

    autofit_columns(ws)
    ws.freeze_panes = "A2"

print(f"\nAll Q18 results saved (formatted) to: {OUTPUT_FILE}")
