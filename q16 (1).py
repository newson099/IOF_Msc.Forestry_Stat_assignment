"""
Q16. Determinant factors for the choice of forest product (Fuel wood /
     Grass / Leaf litter) with respect to socio-economic variables, using
     Multinomial Logistic Regression (MNL) -- appropriate here because
     the dependent variable is NOMINAL (unordered categories), unlike
     the ordinal variables used in Q14/Q15.

     Dependent variable: Multi_Choice_needy_FP
         1 = Fuel wood (reference category)
         2 = Grass
         3 = Leaf litter

"""

import os
import numpy as np
import pandas as pd
from statsmodels.formula.api import mnlogit
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_PATH = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\spss.csv"
OUTPUT_DIR = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\16"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "q16_results.xlsx")

ALPHA = 0.05

df = pd.read_csv(DATA_PATH, encoding="utf-8-sig")
df["District"] = df["District"].replace("2+C182:BC182", "2").astype(int)

df["District_lbl"] = df["District"].map({1: "Bara", 2: "Parsa"})
df["Sex_lbl"] = df["HH_Head_sex"].map({1: "Female", 2: "Male"})
df["Ethnic_lbl"] = df["Ehnicity"].map({1: "Dalit", 2: "Indigenous", 3: "Madhesi", 4: "Brahmin_Chhetri"})
df["Eco_class_lbl"] = df["Eco_class"].map({1: "Poor", 2: "Middle", 3: "Rich"})
df["Occupation_lbl"] = df["Main_Occupation"].map({1: "Agriculture", 2: "Service", 3: "Business"})

fp_labels = {1: "Fuel wood", 2: "Grass", 3: "Leaf litter"}
df["FP_choice_lbl"] = df["Multi_Choice_needy_FP"].map(fp_labels)

print("Forest product choice distribution:")
print(df["FP_choice_lbl"].value_counts(), "\n")

# Fitting the multinomial logit (Fuel wood = reference/base outcome) ----
formula = (
    "Multi_Choice_needy_FP ~ Age + Q('Edu.Hh') + Family_size + LSU + Land_holding_ropani + Dis_from "
    "+ C(District_lbl) + C(Sex_lbl) + C(Ethnic_lbl) + C(Eco_class_lbl) + C(Occupation_lbl)"
)
model = mnlogit(formula, data=df).fit(method="newton", maxiter=100, disp=False)
print(model.summary())

readable_names = {
    "Intercept": "Intercept",
    "C(District_lbl)[T.Parsa]": "District: Parsa (vs Bara)",
    "C(Sex_lbl)[T.Male]": "Sex: Male (vs Female)",
    "C(Ethnic_lbl)[T.Dalit]": "Ethnicity: Dalit (vs Brahmin/Chhetri)",
    "C(Ethnic_lbl)[T.Indigenous]": "Ethnicity: Indigenous (vs Brahmin/Chhetri)",
    "C(Ethnic_lbl)[T.Madhesi]": "Ethnicity: Madhesi (vs Brahmin/Chhetri)",
    "C(Eco_class_lbl)[T.Poor]": "Economic Class: Poor (vs Middle)",
    "C(Eco_class_lbl)[T.Rich]": "Economic Class: Rich (vs Middle)",
    "C(Occupation_lbl)[T.Business]": "Occupation: Business (vs Agriculture)",
    "C(Occupation_lbl)[T.Service]": "Occupation: Service (vs Agriculture)",
    "Age": "Age (years)",
    "Q('Edu.Hh')": "Education (years)",
    "Family_size": "Family Size",
    "LSU": "Livestock Standard Unit",
    "Land_holding_ropani": "Land Holding (ropani)",
    "Dis_from": "Distance from BZ (min)",
}

# Building one clean coefficient table PER outcome category
# statsmodels mnlogit numbers outcomes 0,1,2 internally corresponding to
# the sorted unique values of the dependent variable (1=Fuelwood is the
# base/reference outcome and does not get its own equation).
outcome_map = {0: "Grass (vs Fuel wood)", 1: "Leaf litter (vs Fuel wood)"}

coef_tables = {}
for outcome_idx, outcome_name in outcome_map.items():
    params = model.params[outcome_idx]
    bse = model.bse[outcome_idx]
    tvals = model.tvalues[outcome_idx]
    pvals = model.pvalues[outcome_idx]

    table = pd.DataFrame({
        "Predictor": [readable_names.get(t, t) for t in params.index],
        "Coefficient": params.values.round(4),
        "Std_Error": bse.values.round(4),
        "z_stat": tvals.values.round(3),
        "p_value": pvals.values.round(4),
        "Relative_Risk_Ratio": np.exp(params.values).round(3),
        "Significant_5pct": pvals.values < ALPHA,
    })
    coef_tables[outcome_name] = table
    print(f"\n--- {outcome_name} ---")
    print(table.to_string(index=False))

# Model fitting statistics
pred_class_idx = model.predict(df).values.argmax(axis=1)
sorted_categories = sorted(df["Multi_Choice_needy_FP"].unique())
pred_class = [sorted_categories[i] for i in pred_class_idx]
accuracy = (np.array(pred_class) == df["Multi_Choice_needy_FP"].values).mean()

model_summary = pd.DataFrame({
    "Statistic": ["N", "Log-Likelihood", "LL-Null", "Pseudo R-squared (McFadden)",
                  "LR chi-square", "LR p-value", "Classification Accuracy"],
    "Value": [int(model.nobs), round(model.llf, 3), round(model.llnull, 3),
              round(model.prsquared, 4), round(model.llr, 3), round(model.llr_pvalue, 4),
              f"{accuracy*100:.1f}%"],
})
print("\nModel Summary:")
print(model_summary.to_string(index=False))

fp_dist = df["FP_choice_lbl"].value_counts().reindex(["Fuel wood", "Grass", "Leaf litter"]).reset_index()
fp_dist.columns = ["Forest Product Choice", "N"]

#output
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SIG_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4


def write_table(ws, df_table, start_row=1, title=None, highlight_sig_col=None, first_col_left=False):
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 2

    header_row = row
    for c, col_name in enumerate(df_table.columns, start=1):
        ws.cell(row=header_row, column=c, value=col_name)
    style_header_row(ws, header_row, len(df_table.columns))
    row += 1

    for _, data_row in df_table.iterrows():
        for c, col_name in enumerate(df_table.columns, start=1):
            cell = ws.cell(row=row, column=c, value=data_row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = LEFT if (first_col_left and c == 1) else CENTER
            if highlight_sig_col is not None and col_name == highlight_sig_col and bool(data_row[col_name]):
                for cc in range(1, len(df_table.columns) + 1):
                    ws.cell(row=row, column=cc).fill = SIG_FILL
        row += 1
    return row + 2


with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Q16_Multinomial_Logit")
    wb = writer.book
    ws = wb["Q16_Multinomial_Logit"]

    next_row = write_table(ws, fp_dist, start_row=1, title="Forest Product Choice Distribution")

    next_row = write_table(
        ws, coef_tables["Grass (vs Fuel wood)"], start_row=next_row,
        title="MNL Coefficients: Grass vs Fuel wood (base) -- green = significant at 5%",
        highlight_sig_col="Significant_5pct", first_col_left=True,
    )
    next_row = write_table(
        ws, coef_tables["Leaf litter (vs Fuel wood)"], start_row=next_row,
        title="MNL Coefficients: Leaf litter vs Fuel wood (base) -- green = significant at 5%",
        highlight_sig_col="Significant_5pct", first_col_left=True,
    )
    write_table(ws, model_summary, start_row=next_row, title="Model Fit Summary", first_col_left=True)

    autofit_columns(ws)
    ws.freeze_panes = "A2"

print(f"\nAll Q16 results saved (formatted) to: {OUTPUT_FILE}")
