"""
Q19. Influencing factors for household involvement in BZ activities,
     using Discriminant Function Analysis (DFA).

     Grouping variable: Invo_men (0 = Not Involved, 1 = Involved)
     Predictors: same socio-economic set used in Q11/Q13 (continuous +
                 dummy-coded categorical variables)

     DFA finds the linear combination of predictors ("discriminant
     function") that best separates the two involvement groups, and
     identifies which variables contribute most to that separation.

"""

import os
import numpy as np
import pandas as pd
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from statsmodels.multivariate.manova import MANOVA
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_PATH = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\spss.csv"
OUTPUT_DIR = r"E:\IOF\!st yr 2nd sem\stat\001_Internal_assessment\ass\19"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "q19_results.xlsx")

df = pd.read_csv(DATA_PATH, encoding="utf-8-sig")
df["District"] = df["District"].replace("2+C182:BC182", "2").astype(int)

df["District_lbl"] = df["District"].map({1: "Bara", 2: "Parsa"})
df["Sex_lbl"] = df["HH_Head_sex"].map({1: "Female", 2: "Male"})
df["Ethnic_lbl"] = df["Ehnicity"].map({1: "Dalit", 2: "Indigenous", 3: "Madhesi", 4: "Brahmin_Chhetri"})
df["Eco_class_lbl"] = df["Eco_class"].map({1: "Poor", 2: "Middle", 3: "Rich"})
df["Occupation_lbl"] = df["Main_Occupation"].map({1: "Agriculture", 2: "Service", 3: "Business"})

X = pd.get_dummies(
    df[["Age", "Edu.Hh", "Family_size", "LSU", "Land_holding_ropani", "Dis_from",
        "District_lbl", "Sex_lbl", "Ethnic_lbl", "Eco_class_lbl", "Occupation_lbl"]],
    columns=["District_lbl", "Sex_lbl", "Ethnic_lbl", "Eco_class_lbl", "Occupation_lbl"],
    drop_first=True,
).astype(float)
y = df["Invo_men"]

readable_names = {
    "Age": "Age (years)",
    "Edu.Hh": "Education (years)",
    "Family_size": "Family Size",
    "LSU": "Livestock Standard Unit",
    "Land_holding_ropani": "Land Holding (ropani)",
    "Dis_from": "Distance from BZ (min)",
    "District_lbl_Parsa": "District: Parsa (vs Bara)",
    "Sex_lbl_Male": "Sex: Male (vs Female)",
    "Ethnic_lbl_Dalit": "Ethnicity: Dalit (vs Brahmin/Chhetri)",
    "Ethnic_lbl_Indigenous": "Ethnicity: Indigenous (vs Brahmin/Chhetri)",
    "Ethnic_lbl_Madhesi": "Ethnicity: Madhesi (vs Brahmin/Chhetri)",
    "Eco_class_lbl_Poor": "Economic Class: Poor (vs Middle)",
    "Eco_class_lbl_Rich": "Economic Class: Rich (vs Middle)",
    "Occupation_lbl_Service": "Occupation: Service (vs Agriculture)",
    "Occupation_lbl_Business": "Occupation: Business (vs Agriculture)",
}
X.columns = [readable_names.get(c, c) for c in X.columns]

print("Group distribution (Invo_men):")
print(y.value_counts(), "\n")

manova_df = X.copy()
manova_df["Invo_men"] = y.values
dep_vars = " + ".join([f"Q('{c}')" for c in X.columns])
manova = MANOVA.from_formula(f"{dep_vars} ~ Invo_men", data=manova_df)
manova_result = manova.mv_test()

wilks_row = manova_result.results["Invo_men"]["stat"].loc["Wilks' lambda"]
wilks_lambda = wilks_row["Value"]
wilks_F = wilks_row["F Value"]
wilks_p = wilks_row["Pr > F"]

print("Overall Test of Group Differences (Wilks' Lambda, via MANOVA)")
print(f"  Wilks' Lambda = {wilks_lambda:.4f}, F = {wilks_F:.3f}, p-value = {wilks_p:.4f}")
if wilks_p < 0.05:
    print("  -> Reject H0 at 5% level: the two groups differ significantly on the predictor set.\n")
else:
    print("  -> Fail to reject H0 at 5% level.\n")

wilks_table = pd.DataFrame([{
    "Wilks_Lambda": round(wilks_lambda, 4),
    "F_value": round(wilks_F, 3),
    "p_value": round(wilks_p, 4),
    "Significant_5pct": wilks_p < 0.05,
}])

lda = LinearDiscriminantAnalysis()
lda.fit(X, y)

raw_coef = pd.Series(lda.coef_[0], index=X.columns)
intercept = lda.intercept_[0]

pooled_std = X.std(ddof=1)
standardized_coef = raw_coef * pooled_std

discriminant_scores = lda.decision_function(X)
structure_corr = X.apply(lambda col: np.corrcoef(col, discriminant_scores)[0, 1])

coef_table = pd.DataFrame({
    "Predictor": X.columns,
    "Unstandardized_Coefficient": raw_coef.values.round(4),
    "Standardized_Coefficient": standardized_coef.values.round(4),
    "Structure_Correlation": structure_corr.values.round(3),
}).sort_values("Standardized_Coefficient", key=abs, ascending=False)

print("Discriminant Function Coefficients (sorted by relative importance):")
print(coef_table.to_string(index=False))
print(f"\n(Constant / Intercept = {intercept:.4f})\n")

centroid_table = pd.DataFrame({
    "Group": ["Not Involved (0)", "Involved (1)"],
    "Group_Centroid (mean discriminant score)": [
        discriminant_scores[y == 0].mean(),
        discriminant_scores[y == 1].mean(),
    ],
}).round(4)
print("Group Centroids:")
print(centroid_table.to_string(index=False))

pred = lda.predict(X)
accuracy = (pred == y).mean()
confusion = pd.crosstab(y, pred, rownames=["Actual"], colnames=["Predicted"])
print(f"\nClassification Accuracy = {accuracy*100:.1f}%")
print("Confusion Matrix:")
print(confusion, "\n")

model_summary = pd.DataFrame([{
    "N": len(y),
    "Classification_Accuracy": f"{accuracy*100:.1f}%",
    "Wilks_Lambda": round(wilks_lambda, 4),
    "Wilks_F": round(wilks_F, 3),
    "Wilks_p_value": round(wilks_p, 4),
}])

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SIG_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row_num, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 4


def write_table(ws, df_table, start_row=1, title=None, first_col_left=False):
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 2

    header_row = row
    for c, col_name in enumerate(df_table.columns, start=1):
        ws.cell(row=header_row, column=c, value=col_name)
    style_header_row(ws, header_row, len(df_table.columns))
    row += 1

    for _, data_row in df_table.iterrows():
        for c, col_name in enumerate(df_table.columns, start=1):
            cell = ws.cell(row=row, column=c, value=data_row[col_name])
            cell.border = THIN_BORDER
            cell.alignment = LEFT if (first_col_left and c == 1) else CENTER
        row += 1
    return row + 2


with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Q19_Discriminant_Analysis")
    wb = writer.book
    ws = wb["Q19_Discriminant_Analysis"]

    next_row = write_table(ws, wilks_table, start_row=1, title="Overall Significance Test (Wilks' Lambda / MANOVA)")
    next_row = write_table(
        ws, coef_table, start_row=next_row,
        title="Discriminant Function Coefficients (sorted by |standardized coefficient|)",
        first_col_left=True,
    )
    next_row = write_table(ws, centroid_table, start_row=next_row, title="Group Centroids")
    write_table(ws, model_summary, start_row=next_row, title="Model Summary")

    autofit_columns(ws)
    ws.freeze_panes = "A2"

print(f"All Q19 results saved (formatted) to: {OUTPUT_FILE}")
